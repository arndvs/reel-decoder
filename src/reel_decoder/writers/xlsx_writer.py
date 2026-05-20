"""Step 9: Write — append a decoded row to the Swipe Library xlsx.

If swipe-library.xlsx doesn't exist, create it with the same column schema
as the one in the Reel Swipe Library workbook. If it exists, append a row.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from rich.console import Console

from reel_decoder.schema import DecodedReel

console = Console()


HEADERS = [
    ("#", 5),
    ("Date Added", 12),
    ("Link", 32),
    ("Creator", 18),
    ("Niche", 18),
    ("Hook Pattern", 16),
    ("Hook Text", 35),
    ("Hook Visual", 30),
    ("Beat 1 — Text", 24),
    ("Beat 1 — Type", 14),
    ("Beat 1 — Visual", 24),
    ("Beat 2 — Text", 24),
    ("Beat 2 — Type", 14),
    ("Beat 2 — Visual", 24),
    ("Beat 3 — Text", 24),
    ("Beat 3 — Type", 14),
    ("Beat 3 — Visual", 24),
    ("Beat 4 — Text (opt)", 24),
    ("Beat 4 — Type", 14),
    ("Beat 4 — Visual", 24),
    ("Beat 5 — Text (opt)", 24),
    ("Beat 5 — Type", 14),
    ("Beat 5 — Visual", 24),
    ("Beat 6 — Text (opt)", 24),
    ("Beat 6 — Type", 14),
    ("Beat 6 — Visual", 24),
    ("Mechanism Line", 30),
    ("Payoff Visual", 30),
    ("CTA", 20),
    ("Length (s)", 10),
    ("Music Vibe", 14),
    ("Caption Style", 16),
    ("Why It Works", 45),
    ("Stop-Scroll", 12),
    ("Notes", 30),
]

NAVY = "1F2937"
BORDER_GRAY = "D1D5DB"
HEADER_FILL = PatternFill("solid", start_color=NAVY)
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
thin = Side(border_style="thin", color=BORDER_GRAY)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _init_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Swipe Library"

    for col_idx, (header, width) in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 36
    ws.freeze_panes = "C2"

    # Dropdowns matching the Reel Swipe Library workbook
    hook_dv = DataValidation(
        type="list",
        formula1='"Rarity,Contrarian,Stat,Problem/Agitate,Authority,Transformation,Question,Demonstration"',
        allow_blank=True,
    )
    ws.add_data_validation(hook_dv)
    hook_dv.add("F2:F1000")

    beat_dv = DataValidation(
        type="list",
        formula1='"Benefit,Mechanism,Aspiration,Social Proof,Demonstration,Reveal"',
        allow_blank=True,
    )
    ws.add_data_validation(beat_dv)
    for col_letter in ["J", "M", "P", "S", "V", "Y"]:
        beat_dv.add(f"{col_letter}2:{col_letter}1000")

    music_dv = DataValidation(
        type="list",
        formula1='"Ambient,Driving/EDM,Cinematic,Hip-Hop,Pop,None/Silent"',
        allow_blank=True,
    )
    ws.add_data_validation(music_dv)
    music_dv.add("AE2:AE1000")

    caption_dv = DataValidation(
        type="list",
        formula1='"Karaoke,Static cards,Voiceover only,Mixed"',
        allow_blank=True,
    )
    ws.add_data_validation(caption_dv)
    caption_dv.add("AF2:AF1000")

    scroll_dv = DataValidation(
        type="list", formula1='"1,2,3,4,5"', allow_blank=True
    )
    ws.add_data_validation(scroll_dv)
    scroll_dv.add("AH2:AH1000")

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def append_row(decoded: DecodedReel, xlsx_path: Path) -> int:
    """Append a decoded reel as a new row. Returns the row number written.

    Deduplicates on source_path (column C) — if a row with the same path
    already exists, logs a warning and returns the existing row number.
    """
    if not xlsx_path.exists():
        _init_workbook(xlsx_path)

    wb = load_workbook(xlsx_path)
    ws = wb["Swipe Library"]

    # Dedup: check if source_path already exists in column C
    for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
        if row[2].value == decoded.source_path:
            existing_row = row[0].row
            console.log(f"[yellow]write: skipped duplicate — {decoded.source_path} already at row {existing_row}[/yellow]")
            wb.close()
            return existing_row

    # Find first empty row
    row_num = ws.max_row + 1
    if ws.cell(row=row_num - 1, column=3).value is None and row_num > 2:
        # Last row is empty; use it
        row_num -= 1

    values = decoded.to_xlsx_row()
    # Set # to current count of filled rows
    values[0] = sum(1 for r in ws.iter_rows(min_row=2, max_col=3) if r[2].value) + 1

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=val)
        cell.font = BODY_FONT
        cell.alignment = LEFT
        cell.border = BORDER
    ws.row_dimensions[row_num].height = 55

    wb.save(xlsx_path)
    console.log(f"write: appended row {row_num} to {xlsx_path}")
    return row_num
